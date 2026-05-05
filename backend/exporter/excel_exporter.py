import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WARN_THIN   = Side(style="thin", color="FF4D6D")
_WARN_BORDER = Border(left=_WARN_THIN, right=_WARN_THIN, top=_WARN_THIN, bottom=_WARN_THIN)

_C = {
    "topic_bg": "C6EFCE",
    "raw_bg": "FFF2CC",
    "avro_bg": "E2EFDA",
    "detail_bg": "BDD7EE",
    "col_hdr_bg": "2E75B6",
    "col_hdr_fg": "FFFFFF",
    "pk_bg": "FFD966",
    "logical_bg": "FFFFC0",
    "row_odd": "FFFFFF",
    "row_even": "F2F2F2",
    # byte-anomaly warning colours
    "warn_bg": "FFE7EC",   # light red fill
    "warn_fg": "C0002A",   # dark red text
    "warn_hdr_bg": "FF4D6D",   # red header bar
    "warn_hdr_fg": "FFFFFF",
    "warn_border": "FF4D6D",
    "dup_bg": "FFC7CE",   # duplicate table highlight
}

# Single-table layout: Raw (cols 1-8) | Gap (cols 9-10) | AVRO (cols 11-19)
_RAW_END = 8
_AVRO_START = 11
_AVRO_END = 19

# Multi-table horizontal layout constants
_RAW_COLS = 8    # columns per Raw section
_AVRO_COLS = 9    # columns per AVRO section (9 cols now)
_INNER_GAP = 2    # gap between Raw and AVRO within same table
_TABLE_GAP = 2    # gap between different tables
_BLOCK = _RAW_COLS + _INNER_GAP + _AVRO_COLS + _TABLE_GAP  # = 21 cols per table


def _s(ws, row, col, value, bg=None, fg="000000", bold=False,
       align_h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(name="Arial", size=10, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    c.border = _BORDER
    return c


def _sw(ws, row, col, value, bold=False, align_h="left", wrap=True):
    """Write a warning-styled cell (red bg, red border)."""
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", start_color=_C["warn_bg"])
    c.font = Font(name="Arial", size=10, bold=bold, color=_C["warn_fg"])
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    c.border = _WARN_BORDER
    return c


def _merge(ws, row, c1, c2):
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def _write_warning_section(ws, anomalies: list, start_row: int,
                           c2: int = _AVRO_END) -> int:
    """
    เขียน WARNING section รวม ต่อท้าย sheet สำหรับ byte anomaly
    anomalies = [{ table, column_name, source_type, raw_type, detail, file }, ...]
    c2 = last column to span (dynamic for multi-table horizontal layout)
    """
    if not anomalies:
        return start_row

    c1 = 1

    # ── blank gap ─────────────────────────────────────────
    start_row += 1

    # ── header bar ────────────────────────────────────────
    _merge(ws, start_row, c1, c2)
    c = ws.cell(row=start_row, column=1,
                value=f"⚠  WARNING — Byte Conversion Anomaly ({len(anomalies)} คอลัมน์)  ⚠")
    c.fill = PatternFill("solid", start_color=_C["warn_hdr_bg"])
    c.font = Font(name="Arial", size=11, bold=True, color=_C["warn_hdr_fg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _WARN_BORDER
    ws.row_dimensions[start_row].height = 20
    start_row += 1

    # ── sub-header ────────────────────────────────────────
    _merge(ws, start_row, c1, c2)
    _sw(ws, start_row, 1,
        "คอลัมน์ด้านล่างถูกแปลงเป็น byte แต่ type ต้นทางไม่ใช่ decimal-family "
        "— กรุณาตรวจสอบ mapping และแก้ไขจุดที่ผิดพลาด",
        bold=False, wrap=True)
    ws.row_dimensions[start_row].height = 30
    start_row += 1

    # ── column headers ────────────────────────────────────
    warn_headers = ["NO.", "Table", "Column Name", "Source SQL Type",
                    "Raw Type (byte)", "Detail / คำอธิบาย", "File"]
    for ci, h in enumerate(warn_headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.fill = PatternFill("solid", start_color=_C["warn_hdr_bg"])
        cell.font = Font(name="Arial", size=10, bold=True, color=_C["warn_hdr_fg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _WARN_BORDER
    for ci in range(len(warn_headers) + 1, c2 + 1):
        _sw(ws, start_row, ci, "")
    start_row += 1

    # ── data rows ─────────────────────────────────────────
    for i, a in enumerate(anomalies, 1):
        _sw(ws, start_row, 1, i, align_h="center", wrap=False)
        _sw(ws, start_row, 2, a.get("table", ""))
        _sw(ws, start_row, 3, a.get("column_name", ""))
        _sw(ws, start_row, 4, a.get("source_type", ""))
        _sw(ws, start_row, 5, a.get("raw_type", ""))
        _sw(ws, start_row, 6, a.get("detail", ""), wrap=True)
        _sw(ws, start_row, 7, a.get("file", ""))
        for ci in range(8, c2 + 1):
            _sw(ws, start_row, ci, "")
        ws.row_dimensions[start_row].height = 28
        start_row += 1

    return start_row


def _write_raw_section(ws, table_name: str, columns: list,
                       start_row: int, start_col: int = 1, is_dup: bool = False) -> int:
    """
    Section 1 — Raw (SQL Server) เขียนที่ cols start_col .. start_col+7
    NO. | Name | PK or Unique | Max Length | Format | Nullable | Description | Possible Value
    """
    c1, c2 = start_col, start_col + 7

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1, f"TABLE:    {table_name}" + (" [DUPLICATED]" if is_dup else ""),
       bg=_C["dup_bg"] if is_dup else _C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1, "New [SQL Server]",
       bg=_C["raw_bg"], bold=True, align_h="left")
    start_row += 1

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    for ci, h in enumerate(
        ["NO.", "Name", "PK or Unique", "Max Length",
         "Format", "Nullable", "Description", "Possible Value"], 1
    ):
        _s(ws, start_row, c1 + ci - 1, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]
        is_pk = "Y" if col.get("is_pk") else "N"

        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"

        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]

        _s(ws, r, c1+0, i,                          bg=bg)
        _s(ws, r, c1+1, col.get("column_name", ""), bg=bg, align_h="left")
        _s(ws, r, c1+2, is_pk,                      bg=_C["pk_bg"] if is_pk == "Y" else bg)
        _s(ws, r, c1+3, max_len,                    bg=bg)
        _s(ws, r, c1+4, base_type,                  bg=bg)
        _s(ws, r, c1+5, col.get("nullable", ""),    bg=bg)
        _s(ws, r, c1+6, "",                         bg=bg, align_h="left", wrap=True)
        _s(ws, r, c1+7, "",                         bg=bg, align_h="left", wrap=True)
        start_row += 1

    return start_row


def _write_avro_section(ws, table_name: str, columns: list,
                        start_row: int, start_col: int = _AVRO_START, is_dup: bool = False) -> int:
    """
    Section 2 — Confluent (AVRO) เขียนที่ cols start_col .. start_col+8 (9 cols)
    NO. | Name | Position | Kafka Type | Format | Intype format | Inject move/if | Description | Possible Value
    """
    c1, c2 = start_col, start_col + 8   # 9 columns now

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1,
       f"Topic:    {table_name}" + (" [DUPLICATED]" if is_dup else ""),
       bg=_C["dup_bg"] if is_dup else _C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1, "Content [AVRO]",
       bg=_C["avro_bg"], bold=True, align_h="left")
    start_row += 1

    _merge(ws, start_row, c1, c2)
    _s(ws, start_row, c1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    for ci, h in enumerate(
        ["NO.", "Name", "Partition Key",
         "Raw Format type", "Logical Format type", "direct move / logic",
         "Description", "Possible Value"], 1
    ):
        _s(ws, start_row, c1 + ci - 1, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    # 9th col header (blank, padding to maintain width)
    _s(ws, start_row, c1 + 8, "",
       bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]
        is_pk = "Y" if col.get("is_pk") else "N"

        _s(ws, r, c1+0, i,                              bg=bg)
        _s(ws, r, c1+1, col.get("column_name", ""),     bg=bg, align_h="left")
        _s(ws, r, c1+2, is_pk,                          bg=_C["pk_bg"] if is_pk == "Y" else bg)
        _s(ws, r, c1+3, col.get("raw_type", ""),        bg=bg)
        _s(ws, r, c1+4, col.get("logical_type", ""),    bg=_C["logical_bg"])
        _s(ws, r, c1+5, "Direct move",                  bg=bg)
        _s(ws, r, c1+6, "",                             bg=bg, align_h="left", wrap=True)
        _s(ws, r, c1+7, "",                             bg=bg, align_h="left", wrap=True)
        _s(ws, r, c1+8, "",                             bg=bg, align_h="left", wrap=True)
        start_row += 1

    return start_row


def _set_col_widths(ws, col_offsets: list | None = None):
    if col_offsets is None:
        col_offsets = [1]

    raw_widths  = [8, 22, 14, 12, 16, 12, 40, 40]
    avro_widths = [8, 22, 12, 16, 18, 18, 18, 40, 40]  # 9 cols
    gap_w = 4

    for base in col_offsets:
        for i, w in enumerate(raw_widths):
            ws.column_dimensions[get_column_letter(base + i)].width = w
        # Inner gap (base+8, base+9)
        ws.column_dimensions[get_column_letter(base + 8)].width = gap_w
        ws.column_dimensions[get_column_letter(base + 9)].width = gap_w
        # AVRO cols (base+10 .. base+18) — 9 cols
        for i, w in enumerate(avro_widths):
            ws.column_dimensions[get_column_letter(base + 10 + i)].width = w
        # Outer gap between tables (base+19, base+20)
        ws.column_dimensions[get_column_letter(base + 19)].width = gap_w
        ws.column_dimensions[get_column_letter(base + 20)].width = gap_w


def _set_comparison_widths(ws):
    widths = [18, 22, 20, 16, 16, 18, 18, 12, 8, 20, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _set_summary_widths(ws):
    widths = [20, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_type_comparison_sheet(ws, tables: dict,
                                 source_db: str | None = None,
                                 dest_db: str | None = None):
    row = 1
    _merge(ws, row, 1, 12)
    _s(ws, row, 1, "Type Comparison", bg=_C["topic_bg"], bold=True, align_h="left")
    row += 2

    if source_db or dest_db:
        _s(ws, row, 1, "Source DB", bold=True, align_h="left")
        _s(ws, row, 2, source_db or "")
        row += 1
        _s(ws, row, 1, "Destination DB", bold=True, align_h="left")
        _s(ws, row, 2, dest_db or "")
        row += 2

    headers = [
        "Table", "Column", "Source SQL Type", "Raw Type",
        "Logical Type", "Standard Type", "Final Type",
        "Nullable", "PK", "File", "Source DB", "Dest DB"
    ]
    for ci, h in enumerate(headers, 1):
        _s(ws, row, ci, h, bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    row += 1

    for table_name, columns in tables.items():
        for col in columns:
            _s(ws, row, 1, table_name, align_h="left")
            _s(ws, row, 2, col.get("column_name", ""), align_h="left")
            _s(ws, row, 3, col.get("source_sql_type", ""))
            _s(ws, row, 4, col.get("raw_type", ""))
            _s(ws, row, 5, col.get("logical_type", ""))
            _s(ws, row, 6, col.get("standard_type", ""))
            _s(ws, row, 7, col.get("final_type", ""))
            _s(ws, row, 8, col.get("nullable", ""))
            _s(ws, row, 9, "Y" if col.get("is_pk") else "N")
            _s(ws, row, 10, col.get("file", ""), align_h="left")
            _s(ws, row, 11, source_db or "")
            _s(ws, row, 12, dest_db or "")
            row += 1

    _set_comparison_widths(ws)


def _build_summary_sheet(ws, tables: dict,
                         source_db: str | None = None,
                         dest_db: str | None = None,
                         byte_anomalies: dict | None = None):
    row = 1
    _merge(ws, row, 1, 2)
    _s(ws, row, 1, "Export Summary", bg=_C["topic_bg"], bold=True, align_h="left")
    row += 2

    total_columns = sum(len(cols) for cols in tables.values())
    duplicate_tables = [name for name in tables if "__" in name]
    anomaly_count = sum(len(v) for v in (byte_anomalies or {}).values())

    summary_rows = [
        ("Source DB", source_db or ""),
        ("Destination DB", dest_db or ""),
        ("Exported Tables", len(tables)),
        ("Exported Columns", total_columns),
        ("Duplicate table keys", ", ".join(duplicate_tables) if duplicate_tables else "None"),
        ("Byte Anomalies", anomaly_count),
    ]

    for key, value in summary_rows:
        _s(ws, row, 1, key, bold=True, align_h="left")
        _s(ws, row, 2, value, align_h="left")
        row += 1

    _set_summary_widths(ws)


def _build_sheet(ws, table_name: str, columns: list, anomalies: list | None = None):
    """Single-table sheet: Raw (cols 1-8) และ AVRO (cols 11-18) เขียนพร้อมกันที่ row 1"""
    end_row = _write_raw_section(ws, table_name, columns, start_row=1, start_col=1)
    _write_avro_section(ws, table_name, columns, start_row=1, start_col=_AVRO_START)

    if anomalies:
        tagged = [{**a, "table": table_name} for a in anomalies if isinstance(a, dict)]
        _write_warning_section(ws, tagged, start_row=end_row, c2=_AVRO_END)

    _set_col_widths(ws, col_offsets=[1])


def _build_multi_sheet(ws, tables: dict, byte_anomalies: dict | None = None):
    """
    ทุกตารางอยู่ใน sheet เดียว เรียงแนวตั้ง (ลงมาเรื่อยๆ)
    แต่ละตาราง: Raw (cols 1-8) ซ้าย | gap | AVRO (cols 11-18) ขวา
    WARNING รวมกันทั้งหมดไว้ด้านล่างสุด พร้อมชื่อตาราง
    """
    from collections import Counter
    _col_sig = {n: frozenset(c.get("column_name", "") for c in cols)
                for n, cols in tables.items()}
    _sig_count = Counter(_col_sig.values())
    _dup_tables = {n for n, sig in _col_sig.items() if _sig_count[sig] > 1}

    current_row   = 1
    all_anomalies: list = []

    for table_name, columns in tables.items():
        _is_dup = table_name in _dup_tables
        raw_end  = _write_raw_section(ws, table_name, columns,
                                      start_row=current_row, start_col=1, is_dup=_is_dup)
        avro_end = _write_avro_section(ws, table_name, columns,
                                       start_row=current_row, start_col=_AVRO_START, is_dup=_is_dup)
        current_row = max(raw_end, avro_end) + 2   # 2-row gap ระหว่างตาราง

        for a in (byte_anomalies or {}).get(table_name) or []:
            if isinstance(a, dict):
                all_anomalies.append({**a, "table": table_name})

    if all_anomalies:
        _write_warning_section(ws, all_anomalies,
                               start_row=current_row - 1, c2=_AVRO_END)

    _set_col_widths(ws, col_offsets=[1])


# ── Public API ────────────────────────────────────────────────────

def export_confluent_xlsx(tables: dict,
                           byte_anomalies: dict | None = None,
                           source_db: str | None = None,
                           dest_db: str | None = None) -> io.BytesIO:
    """ทุกตารางใน workbook เดียว: Data Dictionary + Type Comparison + Summary"""
    wb = Workbook()

    ws = wb.active
    ws.title = "Data Dictionary"
    _build_multi_sheet(ws, tables, byte_anomalies)

    comparison_ws = wb.create_sheet(title="Type Comparison")
    _build_type_comparison_sheet(comparison_ws, tables,
                                 source_db=source_db, dest_db=dest_db)

    summary_ws = wb.create_sheet(title="Summary")
    _build_summary_sheet(summary_ws, tables,
                         source_db=source_db, dest_db=dest_db,
                         byte_anomalies=byte_anomalies)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_table_xlsx(columns: list, table_name: str = "Sheet1",
                      anomalies: list | None = None,
                      source_db: str | None = None,
                      dest_db: str | None = None) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    _build_sheet(ws, table_name, columns, anomalies=anomalies)

    comparison_ws = wb.create_sheet(title="Type Comparison")
    _build_type_comparison_sheet(comparison_ws, {table_name: columns},
                                 source_db=source_db, dest_db=dest_db)

    summary_ws = wb.create_sheet(title="Summary")
    _build_summary_sheet(summary_ws, {table_name: columns},
                         source_db=source_db, dest_db=dest_db,
                         byte_anomalies={table_name: anomalies} if anomalies else {})

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Legacy (ชื่อเดิม — main.py ไม่ต้องแก้) ───────────────────────

def export_all_xlsx(tables: dict) -> io.BytesIO:
    return export_confluent_xlsx(tables)


def _build_csv_rows(table_name: str, columns: list, anomalies: list | None = None) -> list:
    rows = []

    # Raw section
    rows.append([f"TABLE:    {table_name}"])
    rows.append(["New [SQL Server]"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "PK or Unique", "Max Length", "Format", "Nullable", "Description", "Possible Value"])
    for i, col in enumerate(columns, 1):
        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"
        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]
        rows.append([i, col.get("column_name", ""), "Y" if col.get("is_pk") else "N",
                     max_len, base_type, col.get("nullable", ""), "", ""])

    rows.append([])  # blank gap

    # Confluent section (Partition Key จาก is_pk)
    rows.append([f"Topic:    {table_name}"])
    rows.append(["Content [AVRO]"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "Partition Key", "Raw Format type", "Logical Format type",
                 "direct move / logic", "Description", "Possible Value"])
    for i, col in enumerate(columns, 1):
        rows.append([i, col.get("column_name", ""), "Y" if col.get("is_pk") else "N",
                     col.get("raw_type", ""), col.get("logical_type", ""),
                     "Direct move", "", ""])

    return rows


def _build_csv_warning_rows(anomalies_with_table: list) -> list:
    """WARNING section รวม สำหรับ CSV — มีคอลัมน์ Table"""
    if not anomalies_with_table:
        return []
    rows = []
    rows.append([])
    rows.append([f"⚠ WARNING — Byte Conversion Anomaly ({len(anomalies_with_table)} คอลัมน์) ⚠"])
    rows.append(["คอลัมน์ด้านล่างถูกแปลงเป็น byte แต่ type ต้นทางไม่ใช่ decimal-family "
                 "— กรุณาตรวจสอบ mapping และแก้ไขจุดที่ผิดพลาด"])
    rows.append(["NO.", "Table", "Column Name", "Source SQL Type", "Raw Type (byte)", "Detail / คำอธิบาย", "File"])
    for i, a in enumerate(anomalies_with_table, 1):
        rows.append([
            i,
            a.get("table",        ""),
            a.get("column_name",  ""),
            a.get("source_type",  ""),
            a.get("raw_type",     ""),
            a.get("detail",       ""),
            a.get("file",         ""),
        ])
    return rows


def _csv_bytes(rows: list) -> io.BytesIO:
    import csv
    import codecs
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    output = io.BytesIO(codecs.BOM_UTF8 + buf.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def export_all_csv(tables: dict, byte_anomalies: dict | None = None) -> io.BytesIO:
    all_rows = []
    all_anomalies = []
    first = True
    for table_name, columns in tables.items():
        if not first:
            all_rows.append([])
        all_rows.extend(_build_csv_rows(table_name, columns))
        first = False
        for a in (byte_anomalies or {}).get(table_name) or []:
            if isinstance(a, dict):
                all_anomalies.append({**a, "table": table_name})

    all_rows.extend(_build_csv_warning_rows(all_anomalies))
    return _csv_bytes(all_rows)


def export_table_csv(columns: list, table_name: str = "Sheet1",
                     anomalies: list | None = None) -> io.BytesIO:
    rows = _build_csv_rows(table_name, columns)
    if anomalies:
        tagged = [{**a, "table": table_name} for a in anomalies if isinstance(a, dict)]
        rows.extend(_build_csv_warning_rows(tagged))
    return _csv_bytes(rows)