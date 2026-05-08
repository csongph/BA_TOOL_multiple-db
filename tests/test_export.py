import unittest
from io import BytesIO
from openpyxl import load_workbook

from backend.exporter.excel_exporter import export_confluent_xlsx


class TestExcelExport(unittest.TestCase):
    def test_export_confluent_xlsx_includes_three_sheets(self):
        tables = {
            'users': [
                {
                    'column_name': 'id',
                    'file': 'test.sql',
                    'raw_type': 'int',
                    'logical_type': 'int',
                    'standard_type': 'INT',
                    'final_type': 'INTEGER',
                    'source_sql_type': 'int',
                    'nullable': 'NOT NULL',
                    'is_pk': True,
                }
            ]
        }

        buf = export_confluent_xlsx(
            tables,
            byte_anomalies={},
            source_db='postgresql',
            dest_db='sqlserver',
        )

        wb = load_workbook(filename=BytesIO(buf.getvalue()))
        self.assertEqual(wb.sheetnames, ['Data Dictionary', 'Type Comparison', 'Summary'])

        comparison = wb['Type Comparison']
        self.assertEqual(comparison['A2'].value, 'Source DB')
        self.assertEqual(comparison['B2'].value, 'postgresql')
        self.assertEqual(comparison['A3'].value, 'Destination DB')
        self.assertEqual(comparison['B3'].value, 'sqlserver')
        self.assertEqual(comparison['A5'].value, 'Table')
        self.assertEqual(comparison['B5'].value, 'Column')

        summary = wb['Summary']
        self.assertEqual(summary['A3'].value, 'Source DB')
        self.assertEqual(summary['B3'].value, 'postgresql')
        self.assertEqual(summary['A4'].value, 'Destination DB')
        self.assertEqual(summary['B4'].value, 'sqlserver')


if __name__ == '__main__':
    unittest.main()
